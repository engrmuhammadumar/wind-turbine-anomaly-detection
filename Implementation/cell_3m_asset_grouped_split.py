"""Cell 3M — deterministic, leakage-safe asset-grouped data split.

Run inside the notebook namespace immediately after Cell 3L:

    %run -i cell_3m_asset_grouped_split.py

The cell assigns every modeling-eligible row to train, validation, or test by
physical ``asset_key``. It balances farm, label, row, event, and asset support;
exports split diagnostics; and changes no pre-existing manifest column except
``split_assignment``.
"""

from __future__ import annotations

import math
import re
from pathlib import Path

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd


# -----------------------------------------------------------------------------
# 1. Validate the audited post-3K/post-3L checkpoint
# -----------------------------------------------------------------------------

if "ROW_LABEL_MANIFEST" not in globals():
    raise NameError("Run Cells 3I–3L before Cell 3M.")

REQUIRED_COLUMNS_3M = {
    "farm",
    "asset_id",
    "asset_key",
    "event_key",
    "source_row_index",
    "timestamp_utc",
    "final_label",
    "modeling_eligible",
    "split_assignment",
}

missing_columns_3m = REQUIRED_COLUMNS_3M - set(ROW_LABEL_MANIFEST.columns)
if missing_columns_3m:
    raise ValueError(
        "ROW_LABEL_MANIFEST is missing columns required by Cell 3M: "
        f"{sorted(missing_columns_3m)}"
    )

EXPECTED_ELIGIBLE_ROWS_3M = 213_537
EXPECTED_PHYSICAL_ASSETS_3M = 36
SPLIT_ORDER_3M = ["train", "validation", "test"]
TARGET_FRACTIONS_3M = np.array([0.70, 0.15, 0.15], dtype=float)
RANDOM_SEED_3M = 20260810
RANDOM_CANDIDATES_3M = 30_000

modeling_eligible_mask_3m = (
    ROW_LABEL_MANIFEST["modeling_eligible"].fillna(False).astype(bool)
)
eligible_rows_before_3m = int(modeling_eligible_mask_3m.sum())

if eligible_rows_before_3m != EXPECTED_ELIGIBLE_ROWS_3M:
    raise ValueError(
        "Cell 3M expects the verified post-retention eligible count of "
        f"213,537, but found {eligible_rows_before_3m:,}."
    )

if ROW_LABEL_MANIFEST["split_assignment"].notna().any():
    raise ValueError(
        "A split assignment already exists. Restart from Cell 3K before "
        "running Cell 3M again."
    )

manifest_length_before_3m = len(ROW_LABEL_MANIFEST)
manifest_index_before_3m = ROW_LABEL_MANIFEST.index.copy()
unchanged_columns_3m = [
    column
    for column in ROW_LABEL_MANIFEST.columns
    if column != "split_assignment"
]
manifest_integrity_hash_before_3m = pd.util.hash_pandas_object(
    ROW_LABEL_MANIFEST.loc[:, unchanged_columns_3m],
    index=True,
).to_numpy(copy=True)


# -----------------------------------------------------------------------------
# 2. Normalize an immutable eligible-row working frame
# -----------------------------------------------------------------------------

eligible_3m = ROW_LABEL_MANIFEST.loc[
    modeling_eligible_mask_3m,
    [
        "farm",
        "asset_id",
        "asset_key",
        "event_key",
        "source_row_index",
        "timestamp_utc",
        "final_label",
    ],
].copy()

eligible_3m["farm"] = eligible_3m["farm"].astype("string").str.strip()
eligible_3m["asset_id"] = eligible_3m["asset_id"].astype("string").str.strip()
eligible_3m["asset_key"] = eligible_3m["asset_key"].astype("string").str.strip()
eligible_3m["event_key"] = eligible_3m["event_key"].astype("string").str.strip()
eligible_3m["final_label"] = (
    eligible_3m["final_label"].astype("string").str.strip().str.lower()
)
eligible_3m["timestamp_utc"] = pd.to_datetime(
    eligible_3m["timestamp_utc"],
    errors="coerce",
    utc=True,
)

if len(eligible_3m) != EXPECTED_ELIGIBLE_ROWS_3M:
    raise ValueError("Eligible-row extraction was not row-conserving.")

required_nonmissing_3m = [
    "farm",
    "asset_id",
    "asset_key",
    "event_key",
    "timestamp_utc",
    "final_label",
]
if eligible_3m[required_nonmissing_3m].isna().any().any():
    bad_columns_3m = eligible_3m[required_nonmissing_3m].columns[
        eligible_3m[required_nonmissing_3m].isna().any()
    ].tolist()
    raise ValueError(
        "Eligible rows have missing split-critical values in: "
        f"{bad_columns_3m}"
    )

for text_column_3m in ["farm", "asset_id", "asset_key", "event_key"]:
    if eligible_3m[text_column_3m].eq("").any():
        raise ValueError(
            f"Eligible rows contain a blank {text_column_3m}."
        )

FARMS_3M = sorted(eligible_3m["farm"].astype(str).unique().tolist())
LABELS_3M = ["normal", "anomaly"]

if FARMS_3M != ["A", "B", "C"]:
    raise ValueError(f"Expected farms A, B, and C; found {FARMS_3M}.")

observed_labels_3m = sorted(
    eligible_3m["final_label"].astype(str).unique().tolist()
)
if observed_labels_3m != sorted(LABELS_3M):
    raise ValueError(
        "Expected exactly the normalized labels normal and anomaly; found "
        f"{observed_labels_3m}."
    )

if eligible_3m[["event_key", "source_row_index"]].duplicated().any():
    raise ValueError("Eligible manifest source keys are not unique.")

asset_identity_check_3m = eligible_3m.groupby(
    "asset_key",
    dropna=False,
).agg(
    farms=("farm", "nunique"),
    asset_ids=("asset_id", "nunique"),
)
if not (
    asset_identity_check_3m["farms"].eq(1).all()
    and asset_identity_check_3m["asset_ids"].eq(1).all()
):
    raise ValueError("An asset_key maps to multiple farm/asset identities.")

event_identity_check_3m = eligible_3m.groupby(
    "event_key",
    dropna=False,
).agg(
    assets=("asset_key", "nunique"),
    labels=("final_label", "nunique"),
)
if not (
    event_identity_check_3m["assets"].eq(1).all()
    and event_identity_check_3m["labels"].eq(1).all()
):
    raise ValueError(
        "An eligible event spans multiple assets or normalized labels, so "
        "an asset-only split would not isolate event sources."
    )

physical_asset_count_3m = int(eligible_3m["asset_key"].nunique())
if physical_asset_count_3m != EXPECTED_PHYSICAL_ASSETS_3M:
    raise ValueError(
        "Cell 3L reported 36 eligible physical assets, but Cell 3M found "
        f"{physical_asset_count_3m}."
    )


# -----------------------------------------------------------------------------
# 3. Build one balance profile per indivisible physical asset
# -----------------------------------------------------------------------------

asset_profile_3m = (
    eligible_3m.groupby("asset_key", as_index=False, dropna=False)
    .agg(
        farm=("farm", "first"),
        asset_id=("asset_id", "first"),
        eligible_rows=("event_key", "size"),
        eligible_events=("event_key", "nunique"),
        first_timestamp_utc=("timestamp_utc", "min"),
        last_timestamp_utc=("timestamp_utc", "max"),
    )
    .sort_values("asset_key")
    .reset_index(drop=True)
)

balance_feature_values_3m: list[np.ndarray] = []
balance_feature_names_3m: list[str] = []
balance_feature_types_3m: list[str] = []
balance_feature_weights_3m: list[float] = []
coverage_feature_indices_3m: list[int] = []


def add_balance_feature_3m(name, feature_type, values, weight, coverage=False):
    values = np.asarray(values, dtype=float)
    if len(values) != len(asset_profile_3m):
        raise ValueError(f"Balance feature {name} has the wrong length.")
    balance_feature_names_3m.append(str(name))
    balance_feature_types_3m.append(str(feature_type))
    balance_feature_values_3m.append(values)
    balance_feature_weights_3m.append(float(weight))
    if coverage:
        coverage_feature_indices_3m.append(len(balance_feature_names_3m) - 1)


for farm_3m in FARMS_3M:
    for label_3m in LABELS_3M:
        subset_3m = eligible_3m.loc[
            eligible_3m["farm"].eq(farm_3m)
            & eligible_3m["final_label"].eq(label_3m)
        ]
        row_counts_3m = (
            subset_3m.groupby("asset_key").size().reindex(
                asset_profile_3m["asset_key"], fill_value=0
            )
        )
        event_counts_3m = (
            subset_3m.groupby("asset_key")["event_key"].nunique().reindex(
                asset_profile_3m["asset_key"], fill_value=0
            )
        )

        row_column_3m = f"{label_3m}_rows"
        event_column_3m = f"{label_3m}_events"
        asset_profile_3m.loc[
            asset_profile_3m["farm"].eq(farm_3m), row_column_3m
        ] = row_counts_3m.loc[
            asset_profile_3m.loc[
                asset_profile_3m["farm"].eq(farm_3m), "asset_key"
            ]
        ].to_numpy()
        asset_profile_3m.loc[
            asset_profile_3m["farm"].eq(farm_3m), event_column_3m
        ] = event_counts_3m.loc[
            asset_profile_3m.loc[
                asset_profile_3m["farm"].eq(farm_3m), "asset_key"
            ]
        ].to_numpy()

        add_balance_feature_3m(
            f"farm_label_rows::{farm_3m}::{label_3m}",
            "eligible_rows",
            row_counts_3m.to_numpy(),
            weight=5.0,
        )
        add_balance_feature_3m(
            f"farm_label_events::{farm_3m}::{label_3m}",
            "eligible_events",
            event_counts_3m.to_numpy(),
            weight=4.0,
        )
        support_values_3m = row_counts_3m.gt(0).astype(int).to_numpy()
        add_balance_feature_3m(
            f"farm_label_assets::{farm_3m}::{label_3m}",
            "supporting_assets",
            support_values_3m,
            weight=2.5,
            coverage=int(support_values_3m.sum()) >= len(SPLIT_ORDER_3M),
        )

for farm_3m in FARMS_3M:
    farm_support_3m = asset_profile_3m["farm"].eq(farm_3m).astype(int).to_numpy()
    add_balance_feature_3m(
        f"farm_assets::{farm_3m}",
        "farm_assets",
        farm_support_3m,
        weight=3.0,
        coverage=int(farm_support_3m.sum()) >= len(SPLIT_ORDER_3M),
    )

for label_3m in LABELS_3M:
    label_support_3m = np.zeros(len(asset_profile_3m), dtype=int)
    for farm_3m in FARMS_3M:
        subset_assets_3m = set(
            eligible_3m.loc[
                eligible_3m["farm"].eq(farm_3m)
                & eligible_3m["final_label"].eq(label_3m),
                "asset_key",
            ].astype(str)
        )
        label_support_3m += asset_profile_3m["asset_key"].astype(str).isin(
            subset_assets_3m
        ).astype(int).to_numpy()
    label_support_3m = (label_support_3m > 0).astype(int)
    add_balance_feature_3m(
        f"label_assets::{label_3m}",
        "label_assets",
        label_support_3m,
        weight=3.0,
        coverage=int(label_support_3m.sum()) >= len(SPLIT_ORDER_3M),
    )

balance_matrix_3m = np.column_stack(balance_feature_values_3m).astype(float)
balance_totals_3m = balance_matrix_3m.sum(axis=0)
balance_weights_3m = np.asarray(balance_feature_weights_3m, dtype=float)

if np.any(balance_totals_3m <= 0):
    empty_features_3m = [
        balance_feature_names_3m[index]
        for index in np.flatnonzero(balance_totals_3m <= 0)
    ]
    raise ValueError(f"Empty split-balance features: {empty_features_3m}")


# -----------------------------------------------------------------------------
# 4. Determine exact asset counts and optimize the deterministic assignment
# -----------------------------------------------------------------------------


def target_asset_counts_3m(number_of_assets):
    raw_counts_3m = TARGET_FRACTIONS_3M * int(number_of_assets)
    counts_3m = np.floor(raw_counts_3m).astype(int)
    remaining_3m = int(number_of_assets - counts_3m.sum())
    fractional_3m = raw_counts_3m - counts_3m
    # A tie is awarded to test before validation so the untouched holdout gets
    # the slightly larger of the two small asset allocations.
    tie_priority_3m = {0: 0, 1: 1, 2: 2}
    award_order_3m = sorted(
        range(len(SPLIT_ORDER_3M)),
        key=lambda index: (fractional_3m[index], tie_priority_3m[index]),
        reverse=True,
    )
    for index_3m in award_order_3m[:remaining_3m]:
        counts_3m[index_3m] += 1
    return counts_3m


target_asset_counts_array_3m = target_asset_counts_3m(physical_asset_count_3m)
expected_asset_counts_array_3m = np.array([25, 5, 6], dtype=int)
if not np.array_equal(
    target_asset_counts_array_3m,
    expected_asset_counts_array_3m,
):
    raise ValueError(
        "The 70/15/15 target should allocate 25/5/6 of the 36 assets, "
        f"but produced {target_asset_counts_array_3m.tolist()}."
    )

target_feature_matrix_3m = (
    TARGET_FRACTIONS_3M[:, None] * balance_totals_3m[None, :]
)


def evaluate_assignment_3m(assignment_codes_3m, return_actual=False):
    actual_3m = np.vstack(
        [
            balance_matrix_3m[assignment_codes_3m == split_code_3m].sum(axis=0)
            for split_code_3m in range(len(SPLIT_ORDER_3M))
        ]
    )
    normalized_error_3m = (
        actual_3m - target_feature_matrix_3m
    ) / np.maximum(balance_totals_3m[None, :], 1.0)
    weighted_mse_3m = float(
        np.sum(balance_weights_3m[None, :] * normalized_error_3m**2)
    )
    maximum_deviation_3m = float(np.max(np.abs(normalized_error_3m)))
    missing_required_coverage_3m = int(
        (actual_3m[:, coverage_feature_indices_3m] <= 0).sum()
    )
    score_3m = (
        weighted_mse_3m
        + 0.25 * maximum_deviation_3m**2
        + 1_000.0 * missing_required_coverage_3m
    )
    if return_actual:
        return score_3m, actual_3m, missing_required_coverage_3m
    return score_3m


assignment_template_3m = np.concatenate(
    [
        np.full(count_3m, split_code_3m, dtype=np.int8)
        for split_code_3m, count_3m in enumerate(target_asset_counts_array_3m)
    ]
)

rng_3m = np.random.default_rng(RANDOM_SEED_3M)
best_assignment_codes_3m = None
best_assignment_score_3m = math.inf

for _ in range(RANDOM_CANDIDATES_3M):
    candidate_codes_3m = rng_3m.permutation(assignment_template_3m)
    candidate_score_3m = evaluate_assignment_3m(candidate_codes_3m)
    if candidate_score_3m < best_assignment_score_3m:
        best_assignment_score_3m = candidate_score_3m
        best_assignment_codes_3m = candidate_codes_3m.copy()

if best_assignment_codes_3m is None:
    raise RuntimeError("The deterministic asset-assignment search failed.")

# Pairwise local improvement preserves the exact 25/5/6 asset counts.
local_search_iterations_3m = 0
while True:
    best_swap_3m = None
    best_swap_score_3m = best_assignment_score_3m
    for left_3m in range(len(best_assignment_codes_3m) - 1):
        for right_3m in range(left_3m + 1, len(best_assignment_codes_3m)):
            if (
                best_assignment_codes_3m[left_3m]
                == best_assignment_codes_3m[right_3m]
            ):
                continue
            trial_codes_3m = best_assignment_codes_3m.copy()
            trial_codes_3m[left_3m], trial_codes_3m[right_3m] = (
                trial_codes_3m[right_3m],
                trial_codes_3m[left_3m],
            )
            trial_score_3m = evaluate_assignment_3m(trial_codes_3m)
            if trial_score_3m < best_swap_score_3m - 1e-15:
                best_swap_score_3m = trial_score_3m
                best_swap_3m = (left_3m, right_3m)
    if best_swap_3m is None:
        break
    left_3m, right_3m = best_swap_3m
    best_assignment_codes_3m[left_3m], best_assignment_codes_3m[right_3m] = (
        best_assignment_codes_3m[right_3m],
        best_assignment_codes_3m[left_3m],
    )
    best_assignment_score_3m = best_swap_score_3m
    local_search_iterations_3m += 1
    if local_search_iterations_3m > 200:
        raise RuntimeError("The split local search did not converge.")

(
    best_assignment_score_3m,
    best_feature_actual_3m,
    missing_required_coverage_3m,
) = evaluate_assignment_3m(best_assignment_codes_3m, return_actual=True)

if missing_required_coverage_3m:
    raise ValueError(
        "No feasible split satisfying required farm/label coverage was found. "
        "ROW_LABEL_MANIFEST was not modified."
    )

observed_asset_counts_3m = np.bincount(
    best_assignment_codes_3m,
    minlength=len(SPLIT_ORDER_3M),
)
if not np.array_equal(
    observed_asset_counts_3m,
    target_asset_counts_array_3m,
):
    raise ValueError("Optimized split does not preserve target asset counts.")

asset_profile_3m["split_assignment"] = [
    SPLIT_ORDER_3M[split_code_3m]
    for split_code_3m in best_assignment_codes_3m
]
asset_to_split_3m = asset_profile_3m.set_index("asset_key")[
    "split_assignment"
].to_dict()

eligible_3m["split_assignment"] = eligible_3m["asset_key"].map(
    asset_to_split_3m
)
if eligible_3m["split_assignment"].isna().any():
    raise ValueError("An eligible asset was not assigned to a split.")


# -----------------------------------------------------------------------------
# 5. Pre-mutation leakage and coverage gates
# -----------------------------------------------------------------------------

asset_split_counts_pre_3m = eligible_3m.groupby("asset_key")[
    "split_assignment"
].nunique()
event_split_counts_pre_3m = eligible_3m.groupby("event_key")[
    "split_assignment"
].nunique()

if asset_split_counts_pre_3m.gt(1).any():
    raise ValueError("At least one physical asset crosses proposed splits.")
if event_split_counts_pre_3m.gt(1).any():
    raise ValueError("At least one event source crosses proposed splits.")

pre_split_label_coverage_3m = eligible_3m.groupby("split_assignment")[
    "final_label"
].nunique().reindex(SPLIT_ORDER_3M, fill_value=0)
if not pre_split_label_coverage_3m.eq(len(LABELS_3M)).all():
    raise ValueError("Every split must contain both normal and anomaly rows.")

pre_split_farm_coverage_3m = eligible_3m.groupby("split_assignment")[
    "farm"
].nunique().reindex(SPLIT_ORDER_3M, fill_value=0)
if not pre_split_farm_coverage_3m.eq(len(FARMS_3M)).all():
    raise ValueError("Every split must contain all three farms.")

for farm_3m in FARMS_3M:
    for label_3m in LABELS_3M:
        category_frame_3m = eligible_3m.loc[
            eligible_3m["farm"].eq(farm_3m)
            & eligible_3m["final_label"].eq(label_3m)
        ]
        supporting_assets_3m = int(category_frame_3m["asset_key"].nunique())
        if supporting_assets_3m >= len(SPLIT_ORDER_3M):
            covered_splits_3m = int(
                category_frame_3m["split_assignment"].nunique()
            )
            if covered_splits_3m != len(SPLIT_ORDER_3M):
                raise ValueError(
                    f"Farm {farm_3m}, label {label_3m} has enough asset "
                    "support for all splits but is not represented in all."
                )


# -----------------------------------------------------------------------------
# 6. Apply the split_assignment mutation and verify its exact boundary
# -----------------------------------------------------------------------------

manifest_asset_keys_3m = (
    ROW_LABEL_MANIFEST["asset_key"].astype("string").str.strip()
)
manifest_split_values_3m = manifest_asset_keys_3m.map(asset_to_split_3m)

if manifest_split_values_3m.loc[modeling_eligible_mask_3m].isna().any():
    raise ValueError("Not all eligible manifest rows resolve to an asset split.")

# Convert the previously empty split column to a nullable string column, then
# write assignments only on modeling-eligible rows.
ROW_LABEL_MANIFEST["split_assignment"] = ROW_LABEL_MANIFEST[
    "split_assignment"
].astype("string")
ROW_LABEL_MANIFEST.loc[
    modeling_eligible_mask_3m,
    "split_assignment",
] = manifest_split_values_3m.loc[modeling_eligible_mask_3m].astype("string")

if len(ROW_LABEL_MANIFEST) != manifest_length_before_3m:
    raise ValueError("Cell 3M changed the manifest row count.")
if not ROW_LABEL_MANIFEST.index.equals(manifest_index_before_3m):
    raise ValueError("Cell 3M changed the manifest index.")

manifest_integrity_hash_after_3m = pd.util.hash_pandas_object(
    ROW_LABEL_MANIFEST.loc[:, unchanged_columns_3m],
    index=True,
).to_numpy()
if not np.array_equal(
    manifest_integrity_hash_before_3m,
    manifest_integrity_hash_after_3m,
):
    raise ValueError(
        "A pre-existing manifest value outside split_assignment changed."
    )

if int(ROW_LABEL_MANIFEST["modeling_eligible"].sum()) != (
    EXPECTED_ELIGIBLE_ROWS_3M
):
    raise ValueError("Cell 3M changed manifest eligibility.")

eligible_assignments_3m = ROW_LABEL_MANIFEST.loc[
    modeling_eligible_mask_3m,
    "split_assignment",
]
ineligible_assignments_3m = ROW_LABEL_MANIFEST.loc[
    ~modeling_eligible_mask_3m,
    "split_assignment",
]

if eligible_assignments_3m.isna().any():
    raise ValueError("At least one modeling-eligible row is unassigned.")
if not eligible_assignments_3m.isin(SPLIT_ORDER_3M).all():
    raise ValueError("An eligible row has an invalid split name.")
if ineligible_assignments_3m.notna().any():
    raise ValueError("Cell 3M assigned a modeling-ineligible row.")

assigned_manifest_3m = ROW_LABEL_MANIFEST.loc[
    modeling_eligible_mask_3m,
    ["asset_key", "event_key", "split_assignment"],
].copy()
assets_crossing_splits_3m = int(
    assigned_manifest_3m.groupby("asset_key")["split_assignment"]
    .nunique()
    .gt(1)
    .sum()
)
events_crossing_splits_3m = int(
    assigned_manifest_3m.groupby("event_key")["split_assignment"]
    .nunique()
    .gt(1)
    .sum()
)
if assets_crossing_splits_3m or events_crossing_splits_3m:
    raise ValueError("Post-mutation asset/event leakage was detected.")


# -----------------------------------------------------------------------------
# 7. Create paper-ready split tables and audit objects
# -----------------------------------------------------------------------------

split_category_3m = pd.CategoricalDtype(SPLIT_ORDER_3M, ordered=True)
eligible_3m["split_assignment"] = eligible_3m["split_assignment"].astype(
    split_category_3m
)
asset_profile_3m["split_assignment"] = asset_profile_3m[
    "split_assignment"
].astype(split_category_3m)

for label_3m in LABELS_3M:
    row_column_3m = f"{label_3m}_rows"
    event_column_3m = f"{label_3m}_events"
    if row_column_3m not in asset_profile_3m:
        asset_profile_3m[row_column_3m] = 0
    if event_column_3m not in asset_profile_3m:
        asset_profile_3m[event_column_3m] = 0
    asset_profile_3m[row_column_3m] = (
        asset_profile_3m[row_column_3m].fillna(0).astype(int)
    )
    asset_profile_3m[event_column_3m] = (
        asset_profile_3m[event_column_3m].fillna(0).astype(int)
    )

ASSET_SPLIT_ASSIGNMENT_3M = asset_profile_3m.loc[
    :,
    [
        "farm",
        "asset_id",
        "asset_key",
        "split_assignment",
        "eligible_rows",
        "eligible_events",
        "normal_rows",
        "anomaly_rows",
        "normal_events",
        "anomaly_events",
        "first_timestamp_utc",
        "last_timestamp_utc",
    ],
].sort_values(["split_assignment", "farm", "asset_key"]).reset_index(drop=True)
ASSET_SPLIT_ASSIGNMENT_3M["split_assignment"] = (
    ASSET_SPLIT_ASSIGNMENT_3M["split_assignment"].astype("string")
)

split_summary_records_3m = []
for split_name_3m, target_fraction_3m in zip(
    SPLIT_ORDER_3M,
    TARGET_FRACTIONS_3M,
):
    subset_3m = eligible_3m.loc[
        eligible_3m["split_assignment"].eq(split_name_3m)
    ]
    split_summary_records_3m.append(
        {
            "split_assignment": split_name_3m,
            "target_fraction": target_fraction_3m,
            "eligible_rows": len(subset_3m),
            "eligible_row_fraction": len(subset_3m) / len(eligible_3m),
            "eligible_events": subset_3m["event_key"].nunique(),
            "event_fraction": (
                subset_3m["event_key"].nunique()
                / eligible_3m["event_key"].nunique()
            ),
            "physical_assets": subset_3m["asset_key"].nunique(),
            "asset_fraction": (
                subset_3m["asset_key"].nunique()
                / eligible_3m["asset_key"].nunique()
            ),
            "farms": subset_3m["farm"].nunique(),
            "normal_rows": int(subset_3m["final_label"].eq("normal").sum()),
            "anomaly_rows": int(subset_3m["final_label"].eq("anomaly").sum()),
        }
    )

SPLIT_SUMMARY_3M = pd.DataFrame(split_summary_records_3m)

farm_label_index_3m = pd.MultiIndex.from_product(
    [SPLIT_ORDER_3M, FARMS_3M, LABELS_3M],
    names=["split_assignment", "farm", "final_label"],
)
SPLIT_FARM_LABEL_SUMMARY_3M = (
    eligible_3m.groupby(
        ["split_assignment", "farm", "final_label"],
        observed=True,
    )
    .agg(
        eligible_rows=("event_key", "size"),
        eligible_events=("event_key", "nunique"),
        physical_assets=("asset_key", "nunique"),
    )
    .reindex(farm_label_index_3m, fill_value=0)
    .reset_index()
)
farm_label_totals_3m = SPLIT_FARM_LABEL_SUMMARY_3M.groupby(
    ["farm", "final_label"]
)[["eligible_rows", "eligible_events", "physical_assets"]].transform("sum")
for metric_3m in ["eligible_rows", "eligible_events", "physical_assets"]:
    SPLIT_FARM_LABEL_SUMMARY_3M[f"{metric_3m}_fraction"] = np.divide(
        SPLIT_FARM_LABEL_SUMMARY_3M[metric_3m],
        farm_label_totals_3m[metric_3m],
        out=np.zeros(len(SPLIT_FARM_LABEL_SUMMARY_3M), dtype=float),
        where=farm_label_totals_3m[metric_3m].to_numpy() > 0,
    )

class_index_3m = pd.MultiIndex.from_product(
    [SPLIT_ORDER_3M, LABELS_3M],
    names=["split_assignment", "final_label"],
)
SPLIT_CLASS_BALANCE_3M = (
    eligible_3m.groupby(
        ["split_assignment", "final_label"],
        observed=True,
    )
    .agg(
        eligible_rows=("event_key", "size"),
        eligible_events=("event_key", "nunique"),
        physical_assets=("asset_key", "nunique"),
    )
    .reindex(class_index_3m, fill_value=0)
    .reset_index()
)
split_row_totals_3m = SPLIT_CLASS_BALANCE_3M.groupby("split_assignment")[
    "eligible_rows"
].transform("sum")
SPLIT_CLASS_BALANCE_3M["within_split_row_fraction"] = np.divide(
    SPLIT_CLASS_BALANCE_3M["eligible_rows"],
    split_row_totals_3m,
    out=np.zeros(len(SPLIT_CLASS_BALANCE_3M), dtype=float),
    where=split_row_totals_3m.to_numpy() > 0,
)

optimization_records_3m = []
for split_code_3m, split_name_3m in enumerate(SPLIT_ORDER_3M):
    for feature_index_3m, feature_name_3m in enumerate(
        balance_feature_names_3m
    ):
        total_3m = balance_totals_3m[feature_index_3m]
        actual_3m = best_feature_actual_3m[split_code_3m, feature_index_3m]
        actual_fraction_3m = actual_3m / total_3m
        optimization_records_3m.append(
            {
                "split_assignment": split_name_3m,
                "balance_feature": feature_name_3m,
                "feature_type": balance_feature_types_3m[feature_index_3m],
                "feature_weight": balance_weights_3m[feature_index_3m],
                "target_fraction": TARGET_FRACTIONS_3M[split_code_3m],
                "actual_count": actual_3m,
                "total_count": total_3m,
                "actual_fraction": actual_fraction_3m,
                "absolute_fraction_error": abs(
                    actual_fraction_3m - TARGET_FRACTIONS_3M[split_code_3m]
                ),
                "coverage_required": (
                    feature_index_3m in coverage_feature_indices_3m
                ),
            }
        )

SPLIT_OPTIMIZATION_DIAGNOSTICS_3M = pd.DataFrame(
    optimization_records_3m
)

leakage_audit_records_3m = [
    {
        "check": "eligible_row_conservation",
        "observed": int(eligible_assignments_3m.notna().sum()),
        "required": EXPECTED_ELIGIBLE_ROWS_3M,
        "passed": int(eligible_assignments_3m.notna().sum())
        == EXPECTED_ELIGIBLE_ROWS_3M,
    },
    {
        "check": "unassigned_eligible_rows",
        "observed": int(eligible_assignments_3m.isna().sum()),
        "required": 0,
        "passed": int(eligible_assignments_3m.isna().sum()) == 0,
    },
    {
        "check": "assigned_ineligible_rows",
        "observed": int(ineligible_assignments_3m.notna().sum()),
        "required": 0,
        "passed": int(ineligible_assignments_3m.notna().sum()) == 0,
    },
    {
        "check": "assets_crossing_splits",
        "observed": assets_crossing_splits_3m,
        "required": 0,
        "passed": assets_crossing_splits_3m == 0,
    },
    {
        "check": "events_crossing_splits",
        "observed": events_crossing_splits_3m,
        "required": 0,
        "passed": events_crossing_splits_3m == 0,
    },
    {
        "check": "eligible_labels_per_split_minimum",
        "observed": int(pre_split_label_coverage_3m.min()),
        "required": len(LABELS_3M),
        "passed": int(pre_split_label_coverage_3m.min()) == len(LABELS_3M),
    },
    {
        "check": "farms_per_split_minimum",
        "observed": int(pre_split_farm_coverage_3m.min()),
        "required": len(FARMS_3M),
        "passed": int(pre_split_farm_coverage_3m.min()) == len(FARMS_3M),
    },
    {
        "check": "manifest_rows_deleted",
        "observed": manifest_length_before_3m - len(ROW_LABEL_MANIFEST),
        "required": 0,
        "passed": manifest_length_before_3m == len(ROW_LABEL_MANIFEST),
    },
    {
        "check": "non_split_manifest_values_changed",
        "observed": int(
            np.count_nonzero(
                manifest_integrity_hash_before_3m
                != manifest_integrity_hash_after_3m
            )
        ),
        "required": 0,
        "passed": np.array_equal(
            manifest_integrity_hash_before_3m,
            manifest_integrity_hash_after_3m,
        ),
    },
]
SPLIT_LEAKAGE_AUDIT_3M = pd.DataFrame(leakage_audit_records_3m)

if not SPLIT_LEAKAGE_AUDIT_3M["passed"].all():
    failed_checks_3m = SPLIT_LEAKAGE_AUDIT_3M.loc[
        ~SPLIT_LEAKAGE_AUDIT_3M["passed"], "check"
    ].tolist()
    raise ValueError(f"Split audit failed: {failed_checks_3m}")


# -----------------------------------------------------------------------------
# 8. Export tables, workbook, and publication-quality split figures
# -----------------------------------------------------------------------------

OUTPUT_ROOT_3M = Path("paper_visuals_3l") / "split_diagnostics_3m"
TABLE_DIR_3M = OUTPUT_ROOT_3M / "tables"
PNG_DIR_3M = OUTPUT_ROOT_3M / "figures_png"
PDF_DIR_3M = OUTPUT_ROOT_3M / "figures_pdf"
for directory_3m in [TABLE_DIR_3M, PNG_DIR_3M, PDF_DIR_3M]:
    directory_3m.mkdir(parents=True, exist_ok=True)


def slug_3m(value):
    return re.sub(r"[^A-Za-z0-9]+", "_", str(value)).strip("_").lower()


table_exports_3m = [
    (
        "T3M01",
        "asset_split_assignment",
        ASSET_SPLIT_ASSIGNMENT_3M,
        "Canonical asset-to-split assignment and asset-level support.",
    ),
    (
        "T3M02",
        "split_summary",
        SPLIT_SUMMARY_3M,
        "Overall row, event, and asset support by split.",
    ),
    (
        "T3M03",
        "split_farm_label_summary",
        SPLIT_FARM_LABEL_SUMMARY_3M,
        "Farm- and label-stratified split support.",
    ),
    (
        "T3M04",
        "split_class_balance",
        SPLIT_CLASS_BALANCE_3M,
        "Within-split class composition and independent support.",
    ),
    (
        "T3M05",
        "split_optimization_diagnostics",
        SPLIT_OPTIMIZATION_DIAGNOSTICS_3M,
        "Target-versus-actual balance for every optimization feature.",
    ),
    (
        "T3M06",
        "split_leakage_audit",
        SPLIT_LEAKAGE_AUDIT_3M,
        "Hard row-conservation and leakage checks.",
    ),
]

table_registry_records_3m = []
for table_id_3m, table_name_3m, table_frame_3m, purpose_3m in table_exports_3m:
    csv_path_3m = TABLE_DIR_3M / (
        f"{table_id_3m}_{slug_3m(table_name_3m)}.csv"
    )
    tex_path_3m = TABLE_DIR_3M / (
        f"{table_id_3m}_{slug_3m(table_name_3m)}.tex"
    )
    table_frame_3m.to_csv(csv_path_3m, index=False)
    try:
        table_frame_3m.to_latex(tex_path_3m, index=False, escape=True)
        latex_path_value_3m = str(tex_path_3m)
    except Exception as latex_error_3m:
        latex_path_value_3m = pd.NA
        print(
            f"LaTeX export skipped for {table_id_3m}: "
            f"{latex_error_3m}"
        )
    table_registry_records_3m.append(
        {
            "table_id": table_id_3m,
            "name": table_name_3m,
            "purpose": purpose_3m,
            "rows": len(table_frame_3m),
            "csv_path": str(csv_path_3m),
            "latex_path": latex_path_value_3m,
        }
    )

SPLIT_TABLE_REGISTRY_3M = pd.DataFrame(table_registry_records_3m)
split_table_registry_path_3m = OUTPUT_ROOT_3M / "split_table_registry_3m.csv"
SPLIT_TABLE_REGISTRY_3M.to_csv(split_table_registry_path_3m, index=False)

workbook_path_3m = OUTPUT_ROOT_3M / "split_diagnostics_3m.xlsx"


def excel_safe_frame_3m(frame):
    """Return an Excel-compatible copy without changing source datetimes."""
    clean_3m = frame.copy()
    for column_3m in clean_3m.columns:
        if isinstance(clean_3m[column_3m].dtype, pd.DatetimeTZDtype):
            clean_3m[column_3m] = (
                clean_3m[column_3m]
                .dt.tz_convert("UTC")
                .dt.tz_localize(None)
            )
    return clean_3m


try:
    with pd.ExcelWriter(workbook_path_3m, engine="openpyxl") as writer_3m:
        for table_id_3m, table_name_3m, table_frame_3m, _ in table_exports_3m:
            sheet_name_3m = f"{table_id_3m}_{slug_3m(table_name_3m)}"[:31]
            excel_safe_frame_3m(table_frame_3m).to_excel(
                writer_3m,
                sheet_name=sheet_name_3m,
                index=False,
            )
        SPLIT_TABLE_REGISTRY_3M.to_excel(
            writer_3m,
            sheet_name="table_registry",
            index=False,
        )

        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        for worksheet_3m in writer_3m.book.worksheets:
            worksheet_3m.freeze_panes = "A2"
            worksheet_3m.auto_filter.ref = worksheet_3m.dimensions
            for header_cell_3m in worksheet_3m[1]:
                header_cell_3m.font = Font(bold=True, color="FFFFFF")
                header_cell_3m.fill = PatternFill(
                    fill_type="solid",
                    fgColor="1F4E78",
                )
                header_cell_3m.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
            for column_number_3m, column_cells_3m in enumerate(
                worksheet_3m.columns,
                start=1,
            ):
                maximum_length_3m = max(
                    len(str(cell_3m.value)) if cell_3m.value is not None else 0
                    for cell_3m in column_cells_3m
                )
                worksheet_3m.column_dimensions[
                    get_column_letter(column_number_3m)
                ].width = min(max(maximum_length_3m + 2, 10), 45)
except ImportError as excel_error_3m:
    workbook_path_3m = None
    print(f"Excel workbook export skipped: {excel_error_3m}")

mpl.rcParams.update(
    {
        "figure.dpi": 120,
        "savefig.dpi": 300,
        "font.family": "DejaVu Serif",
        "font.size": 9,
        "axes.titlesize": 10,
        "axes.labelsize": 9,
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
        "legend.fontsize": 8,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "pdf.fonttype": 42,
        "ps.fonttype": 42,
    }
)

SPLIT_COLORS_3M = {
    "train": "#0072B2",
    "validation": "#E69F00",
    "test": "#009E73",
}
LABEL_COLORS_3M = {
    "normal": "#0072B2",
    "anomaly": "#D55E00",
}
figure_registry_records_3m = []


def save_split_figure_3m(figure, figure_id, name, title, claim, source_table):
    base_name_3m = f"{figure_id}_{slug_3m(name)}"
    png_path_3m = PNG_DIR_3M / f"{base_name_3m}.png"
    pdf_path_3m = PDF_DIR_3M / f"{base_name_3m}.pdf"
    figure.savefig(png_path_3m, bbox_inches="tight", facecolor="white")
    figure.savefig(pdf_path_3m, bbox_inches="tight", facecolor="white")
    plt.close(figure)
    figure_registry_records_3m.append(
        {
            "figure_id": figure_id,
            "title": title,
            "claim_supported": claim,
            "source_table": source_table,
            "png_path": str(png_path_3m),
            "pdf_path": str(pdf_path_3m),
        }
    )


# D3M01 — absolute eligible row support by class.
row_pivot_3m = SPLIT_CLASS_BALANCE_3M.pivot(
    index="split_assignment",
    columns="final_label",
    values="eligible_rows",
).reindex(index=SPLIT_ORDER_3M, columns=LABELS_3M, fill_value=0)
fig_3m, ax_3m = plt.subplots(figsize=(6.6, 4.0))
bottom_3m = np.zeros(len(SPLIT_ORDER_3M))
for label_3m in LABELS_3M:
    values_3m = row_pivot_3m[label_3m].to_numpy()
    ax_3m.bar(
        SPLIT_ORDER_3M,
        values_3m,
        bottom=bottom_3m,
        color=LABEL_COLORS_3M[label_3m],
        label=label_3m.capitalize(),
    )
    bottom_3m += values_3m
ax_3m.set_title("Eligible observations in the asset-grouped split")
ax_3m.set_ylabel("Eligible observations")
ax_3m.grid(axis="y", color="#D9D9D9", linewidth=0.6, alpha=0.7)
ax_3m.legend(frameon=False)
save_split_figure_3m(
    fig_3m,
    "D3M01",
    "eligible_rows_by_split_and_class",
    "Eligible observations by split and class",
    "Shows final class support after physical-asset isolation.",
    "SPLIT_CLASS_BALANCE_3M",
)

# D3M02 — normalized class composition.
composition_pivot_3m = SPLIT_CLASS_BALANCE_3M.pivot(
    index="split_assignment",
    columns="final_label",
    values="within_split_row_fraction",
).reindex(index=SPLIT_ORDER_3M, columns=LABELS_3M, fill_value=0)
fig_3m, ax_3m = plt.subplots(figsize=(6.6, 4.0))
bottom_3m = np.zeros(len(SPLIT_ORDER_3M))
for label_3m in LABELS_3M:
    values_3m = 100.0 * composition_pivot_3m[label_3m].to_numpy()
    ax_3m.bar(
        SPLIT_ORDER_3M,
        values_3m,
        bottom=bottom_3m,
        color=LABEL_COLORS_3M[label_3m],
        label=label_3m.capitalize(),
    )
    bottom_3m += values_3m
ax_3m.set_title("Within-split class composition")
ax_3m.set_ylabel("Eligible observations (%)")
ax_3m.set_ylim(0, 100)
ax_3m.grid(axis="y", color="#D9D9D9", linewidth=0.6, alpha=0.7)
ax_3m.legend(frameon=False)
save_split_figure_3m(
    fig_3m,
    "D3M02",
    "within_split_class_composition",
    "Within-split class composition",
    "Checks that class balance remains comparable across grouped splits.",
    "SPLIT_CLASS_BALANCE_3M",
)

# D3M03 — independent support units.
fig_3m, axes_3m = plt.subplots(1, 2, figsize=(9.0, 3.8))
for ax_3m, metric_3m, title_3m in zip(
    axes_3m,
    ["physical_assets", "eligible_events"],
    ["Physical assets", "Independent event files"],
):
    values_3m = SPLIT_SUMMARY_3M.set_index("split_assignment").reindex(
        SPLIT_ORDER_3M
    )[metric_3m]
    ax_3m.bar(
        SPLIT_ORDER_3M,
        values_3m,
        color=[SPLIT_COLORS_3M[name] for name in SPLIT_ORDER_3M],
    )
    ax_3m.set_title(title_3m)
    ax_3m.set_ylabel("Count")
    ax_3m.grid(axis="y", color="#D9D9D9", linewidth=0.6, alpha=0.7)
fig_3m.suptitle("Independent support retained in each split")
fig_3m.tight_layout()
save_split_figure_3m(
    fig_3m,
    "D3M03",
    "independent_asset_and_event_support",
    "Independent asset and event support by split",
    "Reports evaluation support in independent assets and events, not rows alone.",
    "SPLIT_SUMMARY_3M",
)

# D3M04 — farm/label support within each split.
farm_label_plot_3m = SPLIT_FARM_LABEL_SUMMARY_3M.copy()
farm_label_plot_3m["farm_split"] = (
    farm_label_plot_3m["farm"].astype(str)
    + "–"
    + farm_label_plot_3m["split_assignment"].astype(str).str[:3]
)
farm_split_order_3m = [
    f"{farm_3m}–{split_name_3m[:3]}"
    for farm_3m in FARMS_3M
    for split_name_3m in SPLIT_ORDER_3M
]
farm_label_pivot_3m = farm_label_plot_3m.pivot(
    index="farm_split",
    columns="final_label",
    values="eligible_rows",
).reindex(index=farm_split_order_3m, columns=LABELS_3M, fill_value=0)
fig_3m, ax_3m = plt.subplots(figsize=(9.0, 4.2))
bottom_3m = np.zeros(len(farm_split_order_3m))
for label_3m in LABELS_3M:
    values_3m = farm_label_pivot_3m[label_3m].to_numpy()
    ax_3m.bar(
        farm_split_order_3m,
        values_3m,
        bottom=bottom_3m,
        color=LABEL_COLORS_3M[label_3m],
        label=label_3m.capitalize(),
    )
    bottom_3m += values_3m
ax_3m.set_title("Farm- and class-specific support after asset grouping")
ax_3m.set_ylabel("Eligible observations")
ax_3m.tick_params(axis="x", rotation=40)
ax_3m.grid(axis="y", color="#D9D9D9", linewidth=0.6, alpha=0.7)
ax_3m.legend(frameon=False)
save_split_figure_3m(
    fig_3m,
    "D3M04",
    "farm_class_support_by_split",
    "Farm and class support by split",
    "Verifies that every farm and label remains represented in every split.",
    "SPLIT_FARM_LABEL_SUMMARY_3M",
)

# D3M05 — actual versus target shares for the main support units.
share_plot_3m = SPLIT_SUMMARY_3M.loc[
    :,
    [
        "split_assignment",
        "target_fraction",
        "eligible_row_fraction",
        "event_fraction",
        "asset_fraction",
    ],
].copy()
x_3m = np.arange(len(SPLIT_ORDER_3M), dtype=float)
width_3m = 0.20
fig_3m, ax_3m = plt.subplots(figsize=(7.4, 4.2))
for offset_3m, metric_3m, label_3m in [
    (-width_3m, "eligible_row_fraction", "Rows"),
    (0.0, "event_fraction", "Events"),
    (width_3m, "asset_fraction", "Assets"),
]:
    ax_3m.bar(
        x_3m + offset_3m,
        100.0 * share_plot_3m[metric_3m].to_numpy(),
        width=width_3m,
        label=label_3m,
    )
ax_3m.scatter(
    x_3m,
    100.0 * share_plot_3m["target_fraction"].to_numpy(),
    marker="D",
    s=35,
    color="black",
    label="Target",
    zorder=5,
)
ax_3m.set_xticks(x_3m, SPLIT_ORDER_3M)
ax_3m.set_ylabel("Share of complete eligible dataset (%)")
ax_3m.set_title("Actual split support versus the 70/15/15 target")
ax_3m.grid(axis="y", color="#D9D9D9", linewidth=0.6, alpha=0.7)
ax_3m.legend(frameon=False, ncol=4)
save_split_figure_3m(
    fig_3m,
    "D3M05",
    "actual_vs_target_split_share",
    "Actual versus target split shares",
    "Quantifies unavoidable deviations caused by indivisible physical assets.",
    "SPLIT_SUMMARY_3M",
)

SPLIT_FIGURE_REGISTRY_3M = pd.DataFrame(figure_registry_records_3m)
split_figure_registry_path_3m = OUTPUT_ROOT_3M / "split_figure_registry_3m.csv"
SPLIT_FIGURE_REGISTRY_3M.to_csv(split_figure_registry_path_3m, index=False)


# -----------------------------------------------------------------------------
# 9. Compact notebook report
# -----------------------------------------------------------------------------

print("\nAsset-grouped split assignment:")
print(ASSET_SPLIT_ASSIGNMENT_3M.to_string(index=False))

print("\nOverall split summary:")
print(SPLIT_SUMMARY_3M.to_string(index=False))

print("\nFarm- and label-specific split summary:")
print(SPLIT_FARM_LABEL_SUMMARY_3M.to_string(index=False))

print("\nLeakage and conservation audit:")
print(SPLIT_LEAKAGE_AUDIT_3M.to_string(index=False))

print("\nCell 3M completed successfully.")
print("Optimization seed:", RANDOM_SEED_3M)
print("Random candidate assignments evaluated:", RANDOM_CANDIDATES_3M)
print("Local-search improving swaps:", local_search_iterations_3m)
print("Final weighted balance score:", f"{best_assignment_score_3m:.8f}")
print("Eligible rows assigned:", int(eligible_assignments_3m.notna().sum()))
print("Physical assets assigned:", physical_asset_count_3m)
print("Eligible event files assigned:", eligible_3m["event_key"].nunique())
print(
    "Train/validation/test physical assets:",
    "/".join(str(value) for value in observed_asset_counts_3m),
)
print("Assets crossing splits:", assets_crossing_splits_3m)
print("Events crossing splits:", events_crossing_splits_3m)
print("Unassigned eligible rows:", int(eligible_assignments_3m.isna().sum()))
print("Assigned modeling-ineligible rows:", int(ineligible_assignments_3m.notna().sum()))
print("Split tables generated:", len(SPLIT_TABLE_REGISTRY_3M))
print("Split figures generated:", len(SPLIT_FIGURE_REGISTRY_3M))
print("Split table registry:", split_table_registry_path_3m)
print("Split figure registry:", split_figure_registry_path_3m)
if workbook_path_3m is not None:
    print("Split diagnostics workbook:", workbook_path_3m)
print(
    "\nOnly modeling-eligible rows received a split_assignment. No source row, "
    "label, timestamp, eligibility decision, or audit value was changed."
)
print(
    "Keep the test split untouched until model selection, threshold selection, "
    "and all validation-driven decisions are complete."
)
